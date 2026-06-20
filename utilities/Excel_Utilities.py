import openpyxl
from openpyxl import load_workbook

class Excel_utils_class :
    @staticmethod
    def read_data_from_excel(file_path, sheet_name, row_number, column_number):
        excel_file = openpyxl.load_workbook(file_path)
        sheet = excel_file[sheet_name]
        return sheet.cell(row_number, column_number).value

    @staticmethod
    def write_data_to_excel(file_path, sheet_name, row_number, column_number, data):
        excel_file = openpyxl.load_workbook(file_path)
        sheet = excel_file[sheet_name]
        sheet.cell(row_number, column_number).value = data
        excel_file.save(file_path)
        excel_file.close()


    @staticmethod
    def get_max_row_of_excel(file_path, sheet_name,):
        excel_file = openpyxl.load_workbook(file_path)
        sheet = excel_file[sheet_name]
        return sheet.max_row